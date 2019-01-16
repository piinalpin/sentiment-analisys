from flask import render_template, request #buat memanggil .html
from app import app
from openpyxl import load_workbook
import pandas as pd
import os
import numpy as np
import dbmodel as database
import xlrd
import json
import plotly
from nltk.tokenize import word_tokenize
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
from vectorizer import cluster_paragraphs
from random import shuffle
remover= StopWordRemoverFactory().create_stop_word_remover()
stemmer= StemmerFactory().create_stemmer()
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()




@app.route('/upload', methods= ['GET','POST'])
def upload():
    return render_template("upload_data.html")

@app.route('/upload_hasil', methods= ['GET','POST'])
def upload_hasil():
    if request.method == 'POST':
        a=request.files['file'] #variabel untuk nyimpan file
        b=request.form['sheet']

        a.save(os.path.join('app/upload_data', 'DATA.xlsx')) #wadah untuk setiap kita nge-load, hasilnya disimpan disitu
        wb = load_workbook(filename='app/upload_data/DATA.xlsx')
        sheet_ranges = wb[b]
        x = pd.DataFrame(sheet_ranges.values)
        n = x.dropna(axis = 0, how = 'any')
        a1 = n.replace('-','')
        a2 = a1.replace('--','')
        a3 = a2.replace('---','')
        a4 = a3.replace('',np.NaN)
        a4.dropna(inplace=True)
        a4.columns = ['Id Pegawai','Komentar']
        writer = pd.ExcelWriter('app/upload_data/DATA[CLEANING].xlsx', engine='xlsxwriter') #Hasil cleaning di simpan dalam bentuk file excel
        a4.to_excel(writer, sheet_name=b)
        writer.save() # Ini fungsi buat file excelnya

        wb = load_workbook(filename='app/upload_data/DATA[CLEANING].xlsx') # Baru di load lagi
        sheet_ranges = wb[b]
        data = pd.DataFrame(sheet_ranges.values)
        data.drop(data.columns[[0]], axis = 1, inplace=True) #DROP kolom paling depan

    return render_template('hasil_upload.html', tables=[data.to_html(classes='table table-bordered')],sheet_ranges=b)




@app.route('/hasil_select', methods=['GET','POST'])
def hasil_select():
    if request.method == 'POST':
        select1 = request.form["select1"]
        select2 = request.form["select2"]
        selectkolom = request.form["selectkolom"]
        namakolom = request.form["namakolom"]

        wb = load_workbook(filename='app/upload_data/DATA[CLEANING].xlsx')
        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

        row1 = int(select1)
        row2 = int(select2)

        cols = selectkolom.split(",") #memisahkan kolom dengan koma
        cols = list(map(int,cols)) #convert to list
        xname = namakolom.split(",") #memisahkan nama kolom dengan koma
        data = data[row1:row2][cols]
        data.columns = [xname]

        header = {}
        for index, head in enumerate(xname):
            header[str(index)] = head




        dbmodel = database.DBModel() #memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_cleaning_data("Komentar","datanya",data)
        #result_insert_header = dbmodel.insert_header("DataTA","judulnya",header)

    return render_template('hasil_select.html', tables=[data.to_html(classes='table table-bordered')])

@app.route('/dataframe', methods=['GET','POST'])
def Dataframe():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            data_s.append(stopword.remove(data_baru2))

        data = data_s
        shuffle(data)

        cluster_paragraphs(data, num_clusters=2)
        clusters = cluster_paragraphs(data, num_clusters=2)
        data = pd.DataFrame(clusters)

        

    return render_template('dataframe.html', tables=[data.to_html(classes='table table-bordered')])

@app.route('/index', methods=['GET','POST'])
def Index():
    
    return render_template('index.html')    

@app.route('/tokenizing', methods=['GET','POST'])
def tokenizing():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            word_token2 = word_tokenize(data_baru2)
            data_s.append((word_token2))

    
        data = pd.DataFrame(data_s)
        head = []
        for j in data.columns:
            head_string = "T" + str(j)
            head.append(head_string)

        data.columns = head
        dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_tokenisasi_data("DataTA", "Tokenisasi", data)


    return render_template('tokenizing.html', tables=[data.to_html(classes='table table-bordered')])  

@app.route('/text_filtering', methods=['GET','POST'])
def Text_filtering():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        filter = dbmodel.get_data_all("DataTA","Tokenisasi")

        data_x = []
        for x in filter:
            fils = x.values()
            print fils
            isi_fils = fils[:-1]
            
            data_filter = []
            for z in isi_fils:
                if z <> None:
                    b = (z.encode("ascii","ignore"))
                    stop_w = remover.remove(b)
                    if stop_w <> "":
                        data_filter.append(stop_w)
            data_x.append(data_filter)

            data = pd.DataFrame(data_x)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_filtering_data("DataTA", "Filtering", data)

    return render_template('text_filtering.html', tables=[data.to_html(classes='table table-bordered')]) 
@app.route('/text_stemming', methods=['GET','POST'])
def Text_stemming():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        stem = dbmodel.get_data_all("DataTA","Filtering")

        data_stem = []
        for x in stem:
            stemm = x.values()
            isi_stem = stemm[:-1]
            data_stemming = []
            for z in isi_stem:
                if z <> None:
                    b = (z.encode("ascii", "ignore"))
                    stem_w = stemmer.stem(b)
                    if stem_w <> "":
                        data_stemming.append(stem_w)
            data_stem.append(data_stemming)

            data = pd.DataFrame(data_stem)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_stemming_data("DataTA", "Stemming", data)

        return render_template('text_stemming.html', tables=[data.to_html(classes='table table-bordered')])

@app.route('/tokenizing', methods=['GET','POST'])
def Tokenizing():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        token = dbmodel.get_data_all("DataTA","datanya")

        data_s=[]
        for i in token :
            isi = i.values()
            isi_judul = isi[1]
            data_baru2 = isi_judul.lower()
            word_token2 = word_tokenize(data_baru2)
            data_s.append((word_token2))

    
        data = pd.DataFrame(data_s)
        head = []
        for j in data.columns:
            head_string = "T" + str(j)
            head.append(head_string)

        data.columns = head
        dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
        result_insert_table = dbmodel.insert_tokenisasi_data("DataTA", "Tokenisasi", data)


    return render_template('tokenizing.html', tables=[data.to_html(classes='table table-bordered')])


@app.route('/filtering', methods=['GET','POST'])
def Filtering():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        filter = dbmodel.get_data_all("DataTA","Tokenisasi")

        data_x = []
        for x in filter:
            fils = x.values()
            isi_fils = fils[:-1]
            
            data_filter = []
            for z in isi_fils:
                if z <> None:
                    b = (z.encode("ascii","ignore"))
                    stop_w = remover.remove(b)
                    if stop_w <> "":
                        data_filter.append(stop_w)
            data_x.append(data_filter)

            data = pd.DataFrame(data_x)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_filtering_data("DataTA", "Filtering", data)

    return render_template('filtering.html', tables=[data.to_html(classes='table table-bordered')])

@app.route('/stemming', methods=['GET','POST'])
def Stemming():
    if request.method == 'POST':
        dbmodel = database.DBModel()
        stem = dbmodel.get_data_all("DataTA","Filtering")

        data_stem = []
        for x in stem:
            stemm = x.values()
            isi_stem = stemm[:-1]
            data_stemming = []
            for z in isi_stem:
                if z <> None:
                    b = (z.encode("ascii", "ignore"))
                    stem_w = stemmer.stem(b)
                    if stem_w <> "":
                        data_stemming.append(stem_w)
            data_stem.append(data_stemming)

            data = pd.DataFrame(data_stem)
            head = []
            for j in data.columns:
                head_string = "T" + str(j)
                head.append(head_string)

            data.columns = head
            dbmodel = database.DBModel()  # memanggil file model dimodel class DBModel
            result_insert_table = dbmodel.insert_stemming_data("DataTA", "Stemming", data)

        return render_template('stemming.html', tables=[data.to_html(classes='table table-bordered')])

@app.route('/grafik')
def Grafik():
    rng = pd.date_range('1/1/2011', periods=7500, freq='H')
    ts = pd.Series(np.random.randn(len(rng)), index=rng)

    graphs = [
        dict(
            data=[
                dict(
                    x=[1, 2, 3],
                    y=[10, 20, 30],
                    type='scatter',
                    mode='lines',
                    name = 'Ergonomi'
                ),
                 dict(
                    x=[1, 2, 3],
                    y=[10, 25, 35],
                    type='scatter',
                    mode='lines',
                    name = 'Sistem Produksi'
                ),
                 dict(
                    x=[1, 2, 3],
                    y=[10, 25, 35],
                    type='scatter',
                    mode='lines',
                    name = 'Sistem Produksi'
                ),
                 dict(
                    x=[1, 2, 3],
                    y=[10, 25, 35],
                    type='scatter',
                    mode='lines',
                    name = 'Sistem Produksi'
                )

            ],
            layout=dict(
                title='Trend Data Skripsi Teknik Industri'
            )
        ),

        

        dict(
            data=[
                dict(
                    x=ts.index,  # Can use the pandas data structures directly
                    y=ts
                )
            ]
        )
    ]

    # Add "ids" to each of the graphs to pass up to the client
    # for templating
    ids = [format(i) for i, _ in enumerate(graphs)]

    # Convert the figures to JSON
    # PlotlyJSONEncoder appropriately converts pandas, datetime, etc
    # objects to their JSON equivalents
    graphJSON = json.dumps(graphs, cls=plotly.utils.PlotlyJSONEncoder)

    return render_template('grafik.html',
                           ids=ids,
                           graphJSON=graphJSON)
    






